import numpy as np
import pandas as pd
import time
import json
import re
#import regex as re
print("\n Start " + time.ctime())

updateonly = False #Set True to only search for cards which don't have any associated prices (from a previous run of this code)
alphabetical = True #True searches your cards in alphabetical order. False searches by the highest price (from a previous run of this code)
excelpath = r'moxfield export.xlsx' #This code requires a CSV (in this case, changed to .xlsx) from Moxfield. Other card databases have different column names
dictpath = r'prices_dict.json'
scgpath = r'search-export_2024-10-19.csv'
raritypath = r'moxfield_sell_'
rarities = ['m', 'r', 'u', 'c']
selldf = pd.read_excel(excelpath, header=0) # usecols = [i for i in range(12)] #if you don't want to take the previous price columns

# Use this the first time you build run this code, or if you want to reset your past work
prices = {
    "starcitygames.com": {'Acronym': 'SCG'},
    "cardkingdom.com": {'Acronym': 'CKD'},
    "coolstuffinc.com": {'Acronym': 'CSI'},
    "abugames.com": {"Acronym": "ABU"},
    "capefeargames.com": {"Acronym": "CFG"}
}

#Comment this out if you want to start from scratch
with open(dictpath, 'r') as file:
     prices = json.load(file)

if alphabetical:
    cardnames = np.unique(selldf['Name'])
else:
    max_prices = selldf.groupby('Name')['Max Price'].max()
    # Sort the names based on the maximum prices in descending order
    cardnames = max_prices.sort_values(ascending=False).index.tolist()

for site in prices.keys():
    if site != "coostuffinc.com":
        continue
    column_name = prices[site]['Acronym'] + ' Trade Price'
    if column_name != "SCG Trade Price":
        selldf[column_name] = np.zeros((len(selldf), 1))

doubleface = selldf[selldf['Name'].str.contains(' //')]

for idx in doubleface.index:
    selldf.at[idx, 'Name'] = selldf.at[idx, 'Name'].split(' //')[0] #Most sites list double face cards ('Turn // Burn') as the primary name

del doubleface

#Star City Games gives us a CSV of the buylist so we don't have to scrape their website
def SCG_excel():
    selldf[prices["starcitygames.com"]['Acronym'] + ' Trade Price'] = np.zeros((len(selldf), 1))
    buydf = pd.read_csv(scgpath, header=0,)
    droplist = ['Sealed Product', 'Serialized', 'Alpha', 'Beta']
    for set in droplist:
        buydf = buydf.drop(buydf[buydf['set_name'].str.contains(set)].index.tolist(), axis=0)
    
    for i, card in selldf.iterrows():
        foil = 'N'
        if card['Foil'] == 'foil' or card['Foil'] == 'etched':
            foil = 'F'
        versions = buydf[buydf['name'].str.contains(card['Name'])]
        for n, possibility in versions.iterrows():
            lit = possibility['productid'].split('-')
            #print(card['Name'], lit[2][:3].lower(), card['Edition'], lit[3], card['Collector Number'], lit[4][-2], foil)
            if lit[2][:3].lower() == card['Edition'] and int(lit[3].rstrip('a')) == int(card['Collector Number']) and lit[4][-2] == foil:
                #prices['starcitygames.com'][f"{card['Name']}_{card['Collector Number']}_{card['Edition']}_{foil}"] = possibility['trade_price  (as of 10/19/2024, 8:29:36 AM EST)'] #In case you want to save this to a dict instead of the spreadsheet
                selldf.at[i, prices["starcitygames.com"]['Acronym'] + ' Trade Price'] = possibility['trade_price  (as of 10/19/2024, 8:29:36 AM EST)'] #This column name will change when you download the most recent SCG spreadsheet
                # I'm should dynamically call that column name in the future
                break
    selldf.to_csv(excelpath, index=False)

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from bs4 import BeautifulSoup

def mainLoop():
    #Headless option
    options = Options()
    options.add_argument("--headless")
    options.add_argument("--disable-gpu")  # Disable GPU rendering (optional)
    options.add_argument("--no-sandbox")  # Necessary for some Linux environments
    options.add_argument("--disable-dev-shm-usage") # Overcome resource issues in some environments

    #Provide the path to chromedriver.exe you previously downloaded
    chrome_service = Service(r'chromedriver-win64\chromedriver.exe')

    #Main scraping loop
    for site in prices.keys():
        website = prices[site]['Acronym']
        for n, card in enumerate(cardnames):
            if n % 20 == 0 and n != 0:
                #My internet lags when at lot of people in the apartment complex are at home and streaming. This lets us save and recover before we finish the whole list of cards
                print(n)
                with open(dictpath, 'w') as file:
                    json.dump(prices, file)
            urlname = card.replace(",", "%2C").replace(" ", "+").replace("'", "%27") #Converting this to fit with a URL
            print(website, urlname)
            if website == "CKD":
                url = f'https://www.cardkingdom.com/purchasing/mtg_singles?filter%5Bsort%5D=price_desc&filter%5Bsearch%5D=mtg_advanced&filter%5Bname%5D={urlname}&filter%5Bedition%5D=&filter%5Bformat%5D=&filter%5Bfoils%5D=1&filter%5Bsingles%5D=1&filter%5Bprice_op%5D=&filter%5Bprice%5D='
            elif website == "CSI":
                url = f'https://www.coolstuffinc.com/main_selllist.php?name={urlname}&min=&max=&a=1&s=mtg'
            elif website == "ABU":
                url = f'https://abugames.com/buylist/singles?search={urlname}'
            elif website == 'CFG':
                url = f'https://www.capefeargames.com/buylist#%7B%22q%22%3A%22{urlname}%22%7D'
            print(url)
            # Pass the service object to webdriver.Chrome()
            driver = webdriver.Chrome(service=chrome_service, options=options)
            # open a webpage
            driver.get(url)
            if website == "CKD":
                #CK
                prices["cardkingdom.com"][card] = {}
                CK_buyboxselector = "div.col-sm-9.mainListing"
                WebDriverWait(driver, 5).until(
                        EC.presence_of_element_located((By.CSS_SELECTOR, CK_buyboxselector))
                    )
                page_source = driver.page_source

                # Pass it to BeautifulSoup for parsing
                soup = BeautifulSoup(page_source, 'html.parser')
                body = soup.body
                if body.find('p', class_='noResults col-sm-9'):
                    continue
                CK_mainlisting = body.find('div', class_='col-sm-9 mainListing')
                buylist = CK_mainlisting.find_all('div', class_='productItemWrapper productCardWrapper')
                for i, version in enumerate(buylist):
                    try:
                        set_code = version.find('img', class_='card-image')['alt'][:3]
                        collector_number = version.find('div', class_="collectorNumber").text[-3:]
                        if version.find('div', class_='foil'):
                            foil = 'Foil' #make sure this is the same as CSI
                        else:
                            foil = 'Non-Foil' #make sure this is the same as CSI
                        trade_price = float(version.find('div', class_='creditSellPrice').text.lstrip('$'))
                        prices["cardkingdom.com"][card][str(i)] = {'set_code': set_code,
                                                                    'collector_number': collector_number,
                                                                    'foil': foil,
                                                                    'trade_price': trade_price}
                    except:
                        pass
                    
            elif website == "CSI":
                #CSI
                prices["coolstuffinc.com"][card] = {}
                CSI_buyboxselector = "div.buylist-wrapper.buylist-small-list"
                try:
                    WebDriverWait(driver, 5).until(
                        EC.presence_of_element_located((By.CSS_SELECTOR, CSI_buyboxselector))
                    )
                    
                    # Now that the content is loaded, get the page source
                    page_source = driver.page_source

                    # Pass it to BeautifulSoup for parsing
                    soup = BeautifulSoup(page_source, 'html.parser')

                    # Find the desired <div> or element
                    buybox = soup.select_one(CSI_buyboxselector)

                    if buybox:
                        buylist = buybox.find_all('div', class_='buylist-row-wrapper fullImage')

                        # Loop through each found <div>
                        for i, version in enumerate(buylist):
                            buylist_card_data = version.find('ul', class_='buylist-card-data')
                            if buylist_card_data.find('i'):
                                set_code = [code for code in buylist_card_data.find('i').get('class') if code.startswith('mtg')][0][4:] #This needs to be lower case
                                buylist_card_data_li = buylist_card_data.find_all('li')
                                collector_number = buylist_card_data_li[2].text[1:]
                                foil = buylist_card_data_li[3].text
                                buylist_footer_trade_price = version.find('li', class_='buylist-price-credit')
                                trade_price = float(buylist_footer_trade_price.text[1:buylist_footer_trade_price.text.index("C")])
                                prices["coolstuffinc.com"][card][str(i)] = {'set_code': set_code,
                                                                            'collector_number': collector_number,
                                                                            'foil': foil,
                                                                            'trade_price': trade_price}
                                # print(collector_number)
                                # print(foil)
                                # print(set_code)
                                # print(trade_price)
                            
                    else:
                        print('Element not found')
                except:
                    pass  
            
            elif website == "ABU":
                #ABU
                prices[site][card] = {}
                ABU_buyboxselector = "div.tableHere"
                WebDriverWait(driver, 10).until(
                                    EC.presence_of_element_located((By.CSS_SELECTOR, ABU_buyboxselector))
                                )
                page_source = driver.page_source
                # Pass it to BeautifulSoup for parsing
                soup = BeautifulSoup(page_source, 'html.parser')
                body = soup.body
                if body.find('div', class_='col-lg-3 showningLine'):
                    if "0 of 0" in body.find('div', class_='col-lg-3 showningLine').text:
                        continue
                buylist = body.find_all('div', class_="row panel panel-default not-first buylist ng-star-inserted")
                for i, version in enumerate(buylist):
                    nm = version.find_all('div', class_="col-md-2")[1]
                    if nm.find('span', class_='trade') is not None:
                        trade_price = float(nm.find('span', class_='trade').next_sibling[2:].replace(",","")) #stripping off " $" from left
                    else:
                        trade_price = float(0)
                    block = version.find('div', class_='col-md-3 display-title').text
                    if "(Scroll Showcase" in block:
                        continue
                    if "- FOIL -" in block:
                        foil = 'Foil'
                    else:
                        foil = 'Non-Foil'
                    collector_number = re.findall(r'\((.{3})\)', block)
                    if len(collector_number) > 0:
                        collector_number = collector_number[0]
                    else:
                        collector_number = -999
                    set_name = re.findall(r'\s-\s(?:FOIL\s-)?(.*)', block)[0]
                    prices[site][card][str(i)] = {'set_name': set_name,
                                                'collector_number': collector_number,
                                                'foil': foil,
                                                'trade_price': trade_price}
            elif website == "CFG":
                #CFG
                prices[site][card] = {}
                CFG_buyboxselector = "div.buylist-search-ctr.eat-both.column"
                WebDriverWait(driver, 10).until(
                                    EC.presence_of_element_located((By.CSS_SELECTOR, CFG_buyboxselector))
                                )
                page_source = driver.page_source
                # Pass it to BeautifulSoup for parsing
                soup = BeautifulSoup(page_source, 'html.parser')
                body = soup.body
                if body.find('p', class_='no-results'):
                    continue
                buylist_minus_featured = body.find('div', class_='ajax-buylist browse small-12 column')
                buylist = buylist_minus_featured.find_all('li', class_="product")
                for i, version in enumerate(buylist):
                    if version.find('div', class_='variant-row row no-stock'):
                        continue
                    links = version.find_all('a')
                    skip = False
                    for link in links:
                        href = link.get('href')
                        if href and ('yugioh' in href or 'lorcana' in href or 'pokemon' in href):
                            skip = True
                            break
                    if skip:
                        continue
                    name_block = version.find('h4', class_='name small-12 medium-4')
                    true_name = re.search(r'^(.*?)\s(?:-|\()', name_block.text)
                    if true_name and true_name.group(1) != card:
                        continue
                    if "(" in name_block:
                        collector_number = int(re.search(r'\((.*?)\)', name_block).group(1))
                    else:
                        collector_number = -999
                    if ' Foil' in name_block.text:
                        foil = 'Foil'
                    else:
                        foil = 'Non-Foil'
                    set_name = version.find('span', class_='category').text
                    trade_price = version.find('span', class_='store-credit').text
                    trade_price = float(re.search(r'\$(.*)', trade_price).group(1).replace(",","").lstrip())
                    prices[site][card][str(i)] = {'set_name': set_name,
                                                'collector_number': collector_number,
                                                'foil': foil,
                                                'trade_price': trade_price}                
            driver.quit()
        # print(body.prettify())

###execute       
SCG_excel()
mainLoop()

#Save the dictionary as a json so we can access it without scraping again
with open(dictpath, 'w') as file:
    json.dump(prices, file)

##Final Prices bit
for site in prices.keys():
    column_name = prices[site]['Acronym'] + ' Total Price'
    if column_name not in selldf.columns:
        selldf[column_name] = np.zeros((len(selldf), 1))

#Build Set Codes for websites which only had set names
set_df = pd.read_excel(r'MTG Sets and Codes.xlsx')
tempcopy = prices
#making a copy of the dictionary for safety so I'm not modifying it as I loop over it, but maybe that's not necessary
for site in tempcopy.keys():
    if site != 'coolstuffinc.com':
        continue
    for card in tempcopy[site].keys():
        if card == 'Acronym':
            continue
        for version in tempcopy[site][card]:
            result = set_df.loc[set_df['Set Name'] == prices[site][card][version]['set_name'].rstrip(), 'Set Code']
            if not result.empty:
                prices[site][card][version]['set_code'] = result.iloc[0]
            else:
                prices[site][card][version]['set_code'] = 'UNK'
 
# Build out our prices from dictionary:
for i, card in selldf.iterrows():
    if card['Foil'] == 'foil':
        foil = 'Foil'
    else:
        foil = 'Non-Foil'
    for site in prices.keys():
        if site != "coolstuffinc.com":
            continue
        if card['Name'] in prices[site].keys() and isinstance(prices[site][card['Name']], dict):
            if card['Name'] == "Avacyn, Angel of Hope":
                pass
            for version in prices[site][card['Name']]:
                if prices[site][card['Name']][version]['foil'] == foil and prices[site][card['Name']][version]['set_code'].lower() == card['Edition'] and int(prices[site][card['Name']][version]['collector_number']) == int(card['Collector Number']):
                    selldf.at[i, prices[site]['Acronym'] + ' Trade Price'] = prices[site][card['Name']][version]['trade_price']
                    break
                elif (site == 'abugames.com' or site == 'capefeargames.com') and prices[site][card['Name']][version]['foil'] == foil and prices[site][card['Name']][version]['set_code'].lower() == card['Edition'] and prices[site][card['Name']][version]['collector_number'] == -999:
                    #to handle that ABU and CFG only has a collector number if there are multiple versions in the same set
                    selldf.at[i, prices[site]['Acronym'] + ' Trade Price'] = prices[site][card['Name']][version]['trade_price']
                    break

maxlist = []
for site in prices.keys():
    selldf[prices[site]['Acronym'] + ' Total Price'] = selldf['Tradelist Count'] * selldf[prices[site]['Acronym'] + ' Trade Price']
    maxlist.append(prices[site]['Acronym'] + ' Total Price')


price_columns = selldf[maxlist]
selldf['Max Price'] = price_columns.max(axis=1)
selldf['Best Seller'] = price_columns.idxmax(axis=1) 
selldf.loc[(price_columns == 0).all(axis=1), 'Best Seller'] = 'None'

selldf.to_excel(excelpath, index=False)

def setrarities():
    for rarity in rarities:
        raritydf = pd.read_csv(raritypath + rarity + ".csv", header=0)
        doubleface = raritydf[raritydf['Name'].str.contains(' //')]
        for idx in doubleface.index:
            raritydf.at[idx, 'Name'] = raritydf.at[idx, 'Name'].split(' //')[0]
        columns_to_match = ['Name', 'Edition', 'Foil', 'Collector Number']
        # Perform an inner merge to get the matching rows
        for col in columns_to_match:
            selldf[col] = selldf[col].astype(str)
            raritydf[col] = raritydf[col].astype(str)
        matched_rows = selldf.merge(raritydf[columns_to_match], on=columns_to_match, how='inner')
        # Create a mask by checking if each row in selldf is in the matched rows
        mask = selldf[columns_to_match].isin(matched_rows[columns_to_match].to_dict(orient='list')).all(axis=1)
        selldf.loc[mask, 'Rarity'] = rarity
    selldf.to_csv(excelpath, index=False)


setrarities()

def setcardcolor():
    colors = ['White', 'Blue', 'Black', 'Red', 'Green', 'Gold', 'Colorless']
    for rarity in rarities:
        raritydf = pd.read_csv(raritypath + rarity + ".csv", header=0)
        doubleface = raritydf[raritydf['Name'].str.contains(' //')]
        for idx in doubleface.index:
            raritydf.at[idx, 'Name'] = raritydf.at[idx, 'Name'].split(' //')[0]
        mask = selldf.set_index(['Name', 'Edition', 'Foil', 'Collector Number']).index.isin(
            raritydf.set_index(['Name', 'Edition', 'Foil', 'Collector Number']).index)
        selldf.loc[mask, 'Rarity'] = rarity
    selldf.to_csv(excelpath, index=False)

setcardcolor()

#Bulk opportunity costs from CSI. Feel free to change these if you have a different store which is local or easier to ship to. Assumes Near Mint.
bulk = {'MF': 0.44,
'MN': 0.31,
'RF': 0.13,
'RN': 0.08,
'UF':  0.01,
'UN': 0.00625,
'CF': 0.01,
'CN': 0.00625,
'PN': 0.06} 

def markbulk():
    for combo in bulk.keys():
        if combo[1] == 'F':
            foil = 'foil'
            mask = (selldf['Rarity'] == combo[0].lower()) & (selldf['Foil'] == foil) & (selldf['Max Price'] <= bulk[combo]) & (selldf['Max Price'] > 0)
        else:
            mask = (selldf['Rarity'] == combo[0].lower()) & selldf['Foil'].isna() & (selldf['Max Price']/selldf['Tradelist Count'] <= bulk[combo]) & (selldf['Max Price'] > 0)
        selldf.loc[mask, 'Best Seller'] = 'Bulk - Price Below'
        #print("holder")
        selldf.loc[selldf['Best Price'] == 0, 'Best Location'] = 'Bulk - No Price'
    selldf.to_csv(excelpath, index=False)

markbulk()

def opportunitycostoflessshipping(maxlist):
    max_sum = selldf[maxlist].max(axis=1).sum()
    sum_series = selldf[maxlist].sum(axis=0)
    sorted_sum_series = sum_series[sum_series.argsort()[::-1]]
    reordered_series = pd.concat([
    sorted_sum_series[['CSI Total Price']],
    sorted_sum_series.drop('CSI Total Price')
        ])
    print(reordered_series)

    # Difference between max of all columns and max of subsequent -1 columns
    value_perms = [("All", max_sum)]

    while len(reordered_series) > 1:
        local = reordered_series.index[-1]
        reordered_series = reordered_series.iloc[:-1]  # Remove the last column
        max_remaining = selldf[list(reordered_series.index)].max(axis=1).sum()
        value_perms.append((local, max_remaining))

    # Displaying results
    for i in range(1, len(value_perms)):
        print("Loss without " + value_perms[i][0] + " is $" + str(value_perms[i-1][1] - value_perms[i][1]))

opportunitycostoflessshipping(maxlist)

def valuepersite():
    for site in prices.keys():
        column_name = prices[site]['Acronym'] + ' Total Price'
        thisseller =selldf.loc[selldf['Best Seller'] == column_name]
        total = thisseller['Max Price'].sum()
        quantity = thisseller['Tradelist Count'].sum()
        print(site, "$"+ str(total), "Cards: " + str(quantity))

valuepersite()

##Make a colored version of this sheet

import openpyxl
from openpyxl.styles import PatternFill

# Define fill colors for each column
fill_colors = {
    'SCG Total Price': PatternFill(start_color='B53737', end_color='B53737', fill_type='solid'), # Red
    'CKD Total Price': PatternFill(start_color='C8A2C8', end_color='C8A2C8', fill_type='solid'), # Lilac
    'CSI Total Price': PatternFill(start_color='99CCFF', end_color='99CCFF', fill_type='solid'), # Blue
    'ABU Total Price': PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid'), # Orange
    'CFG Total Price': PatternFill(start_color='99FF99', end_color='99FF99', fill_type='solid'), # Green
}

# Define the color mappings for each code
color_map = {
    'w': 'FFFFFF',   # WHite
    'u': 'ADD8E6',   # Blue
    'b': '838383',   # Dark Gray
    'r': 'FF0000',   # Red
    'g': '008000',   # Green
    'c': 'D3D3D3',   # Light Gray
    'm': 'FFA500'    # Orange
}


trade_columns = [col.replace("Total", "Trade") for col in maxlist]
# Save the dataframe to an Excel file with openpyxl engine
output_file = excelpath[:-5]+'_colored.xlsx'

# Write the dataframe to an Excel file
with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    selldf.to_excel(writer, index=False, sheet_name='Sheet1')
    workbook = writer.book
    worksheet = writer.sheets['Sheet1']

    # Loop through rows and color-code based on the 'Largest Column' value
    for row_idx, row in selldf.iterrows():
        color_key = selldf.at[row_idx, 'Color']
        fill_color = color_map.get(color_key, 'FFFFFF')  # Default to white if color not in map
        fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        worksheet.cell(row=row_idx + 2, column=3).fill = fill  # Apply fill to 'Name' column
        largest_col = row['Best Seller']
        if largest_col in fill_colors:
            # Find the column index to apply the fill
            col_idx = selldf.columns.get_loc(largest_col) + 1  # +1 to account for Excel's 1-based index
            worksheet.cell(row=row_idx + 2, column=col_idx).fill = fill_colors[largest_col]  # row_idx+2 to adjust for header
            # Color the corresponding value in 'Largest Column' with the same color
            largest_col_idx = selldf.columns.get_loc('Best Seller') + 1
            worksheet.cell(row=row_idx + 2, column=largest_col_idx).fill = fill_colors[largest_col]
         # Check if replacing "Total" with "Trade" yields a valid column
        if 'Total' in largest_col:
            trade_col = largest_col.replace('Total', 'Trade')
            if trade_col in selldf.columns:
                # Find the column index for the 'Trade' equivalent
                trade_col_idx = selldf.columns.get_loc(trade_col) + 1  # +1 for Excel's 1-based index
                worksheet.cell(row=row_idx + 2, column=trade_col_idx).fill = fill_colors[largest_col]  # Apply the same color

##File is now saved with color coding

print("exit " + time.ctime())